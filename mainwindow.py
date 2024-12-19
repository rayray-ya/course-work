from PySide6.QtCore import (QCoreApplication, QDate, QLocale,
                            QMetaObject, QObject, QPoint, QRect,
                            QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
                           QFont, QFontDatabase, QGradient, QIcon,
                           QImage, QKeySequence, QLinearGradient, QPainter,
                           QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton,
                               QLabel, QLineEdit, QMessageBox, QComboBox, QCalendarWidget,
                               QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView,
                               QTimeEdit, QGridLayout, QDateEdit)
from database import Database
from purchase_window import PurchaseDialog
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        if not MainWindow.objectName():
            MainWindow.setObjectName(u"MainWindow")
        MainWindow.resize(900, 500)
        MainWindow.setMinimumSize(QSize(900, 500))
        MainWindow.setMaximumSize(QSize(1920, 1080))

        # Set default font size for the entire window
        default_font = QFont()
        default_font.setPointSize(20)
        MainWindow.setFont(default_font)

        # Set the window icon
        icon = QIcon("pineapple.ico")
        MainWindow.setWindowIcon(icon)
        MainWindow.setMouseTracking(True)
        MainWindow.setStyleSheet(u"background-color: #f0f0f0;")

        self.centralwidget = QWidget(MainWindow)
        self.centralwidget.setObjectName(u"centralwidget")

        # Create main layout
        self.mainLayout = QVBoxLayout(self.centralwidget)

        # Title
        self.titleLabel = QLabel("AnAvia", self.centralwidget)
        self.titleLabel.setStyleSheet("font-size: 36px; font-weight: bold; color: white; margin: 10px;")
        self.titleLabel.setAlignment(Qt.AlignCenter)
        self.mainLayout.addWidget(self.titleLabel)

        # Search parameters section
        self.searchLayout = QVBoxLayout()

        # Cities
        self.citiesLayout = QHBoxLayout()
        self.fromLabel = QLabel("Откуда:", self.centralwidget)
        self.fromCity = QLineEdit(self.centralwidget)
        self.toLabel = QLabel("Куда:", self.centralwidget)
        self.toCity = QLineEdit(self.centralwidget)

        self.citiesLayout.addWidget(self.fromLabel)
        self.citiesLayout.addWidget(self.fromCity)
        self.citiesLayout.addWidget(self.toLabel)
        self.citiesLayout.addWidget(self.toCity)

        # Dates
        self.datesLayout = QHBoxLayout()
        self.departLabel = QLabel("Дата вылета:", self.centralwidget)
        self.departDate = QDateEdit(self.centralwidget)
        self.departDate.setCalendarPopup(True)
        self.returnLabel = QLabel("Дата возврата:", self.centralwidget)
        self.returnDate = QDateEdit(self.centralwidget)
        self.returnDate.setCalendarPopup(True)

        self.datesLayout.addWidget(self.departLabel)
        self.datesLayout.addWidget(self.departDate)
        self.datesLayout.addWidget(self.returnLabel)
        self.datesLayout.addWidget(self.returnDate)

        # Passengers and class
        self.detailsLayout = QHBoxLayout()
        self.passengersLabel = QLabel("Пассажиры:", self.centralwidget)
        self.passengersCount = QComboBox(self.centralwidget)
        for i in range(1, 10):
            self.passengersCount.addItem(str(i))

        self.classLabel = QLabel("Класс:", self.centralwidget)
        self.travelClass = QComboBox(self.centralwidget)
        self.travelClass.addItems(["Эконом", "Бизнес", "Первый"])

        self.detailsLayout.addWidget(self.passengersLabel)
        self.detailsLayout.addWidget(self.passengersCount)
        self.detailsLayout.addWidget(self.classLabel)
        self.detailsLayout.addWidget(self.travelClass)

        # Search button
        self.searchButton = QPushButton("Найти билеты", self.centralwidget)

        # Add all layouts to main search layout
        self.searchLayout.addLayout(self.citiesLayout)
        self.searchLayout.addLayout(self.datesLayout)
        self.searchLayout.addLayout(self.detailsLayout)
        self.searchLayout.addWidget(self.searchButton)

        # Results section
        self.resultsLabel = QLabel("Результаты поиска:", self.centralwidget)
        self.resultsLabel.hide()
        self.mainLayout.addWidget(self.resultsLabel)

        # Results table
        self.resultsTable = QTableWidget(self.centralwidget)
        self.resultsTable.setColumnCount(7)
        self.resultsTable.setHorizontalHeaderLabels([
            "Авиакомпания", "Откуда", "Куда", "Дата вылета", "Время вылета", "Время прилета", "Цена"
        ])
        self.resultsTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.resultsTable.hide()
        self.mainLayout.addStretch()
        self.mainLayout.addWidget(self.resultsTable)

        # Add search layout to main layout
        self.mainLayout.addLayout(self.searchLayout)

        # Set central widget
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(QCoreApplication.translate("MainWindow", u"AnAvia", None))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.db = Database()
        self.current_role = "user"

        # Set window title
        self.setWindowTitle("AnAvia")

        # Connect signals
        self.ui.searchButton.clicked.connect(self.search_tickets)

        # Set current date as minimum
        current_date = QDate.currentDate()
        self.ui.departDate.setMinimumDate(current_date)
        self.ui.returnDate.setMinimumDate(current_date)

        # Apply styling
        self.apply_style()

        # Initialize admin controls (but don't show them yet)
        self.setup_admin_controls()

        # Add export button
        self.export_button = QPushButton("Экспорт билетов в Excel", self)
        self.export_button.clicked.connect(self.export_tickets_to_excel)
        self.ui.mainLayout.addWidget(self.export_button)

    def apply_style(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #2b5876;
            }
            QWidget {
                background-color: #2b5876;
            }
            QLabel {
                color: white;
                font-size: 14px;
                padding: 5px;
            }
            QLineEdit {
                padding: 8px;
                border: 2px solid #4e4376;
                border-radius: 4px;
                background-color: white;
                color: black;
                font-size: 14px;
            }
            QPushButton {
                background-color: #4e4376;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 14px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #5e5386;
            }
            QComboBox {
                padding: 8px;
                border: 2px solid #4e4376;
                border-radius: 4px;
                background-color: white;
                color: black;
                font-size: 14px;
                min-width: 150px;
            }
            QDateEdit {
                padding: 8px;
                border: 2px solid #4e4376;
                border-radius: 4px;
                background-color: white;
                color: black;
                font-size: 14px;
                min-width: 150px;
            }
            QTableWidget {
                background-color: white;
                color: black;
                border: 2px solid #4e4376;
                border-radius: 4px;
                font-size: 14px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #4e4376;
                color: white;
                padding: 8px;
                border: none;
                font-size: 14px;
            }
            QSpinBox {
                padding: 8px;
                border: 2px solid #4e4376;
                border-radius: 4px;
                background-color: white;
                color: black;
                font-size: 14px;
                min-width: 100px;
            }
        """)

    def setup_admin_controls(self):
        # Create admin section
        self.admin_widget = QWidget()
        self.admin_widget.setObjectName("adminPanel")
        admin_layout = QVBoxLayout()

        # Add ticket section label
        admin_label = QLabel("Панель администратора")
        admin_label.setStyleSheet("font-size: 16px; font-weight: bold; color: white;")
        admin_layout.addWidget(admin_label)

        # Add fields for new ticket
        ticket_form = QGridLayout()

        # Airline input
        self.ui.airlineInput = QLineEdit()
        self.ui.airlineInput.setPlaceholderText("Название авиакомпании")
        ticket_form.addWidget(QLabel("Авиакомпания:"), 0, 0)
        ticket_form.addWidget(self.ui.airlineInput, 0, 1)

        # Price input
        self.ui.priceInput = QLineEdit()
        self.ui.priceInput.setPlaceholderText("Цена")
        ticket_form.addWidget(QLabel("Цена:"), 1, 0)
        ticket_form.addWidget(self.ui.priceInput, 1, 1)

        # Time inputs
        self.ui.departureTime = QTimeEdit()
        self.ui.arrivalTime = QTimeEdit()
        ticket_form.addWidget(QLabel("Время вылета:"), 2, 0)
        ticket_form.addWidget(self.ui.departureTime, 2, 1)
        ticket_form.addWidget(QLabel("Время прилета:"), 3, 0)
        ticket_form.addWidget(self.ui.arrivalTime, 3, 1)

        admin_layout.addLayout(ticket_form)

        # Add ticket button
        self.ui.addTicketButton = QPushButton("Добавить билет")
        self.ui.addTicketButton.clicked.connect(self.add_ticket)
        admin_layout.addWidget(self.ui.addTicketButton)

        self.admin_widget.setLayout(admin_layout)
        self.ui.mainLayout.addWidget(self.admin_widget)
        self.admin_widget.hide()  # Hide by default

    def hide_admin_controls(self):
        if hasattr(self, 'admin_widget'):
            self.admin_widget.hide()

    def show_admin_controls(self):
        if hasattr(self, 'admin_widget'):
            self.admin_widget.show()

    def set_user(self, login, role):
        """Set the current user and update UI accordingly"""
        self.current_user = login
        self.current_role = role
        self.user_id = self.db.get_user_id(login)  # Добавим получение user_id
        
        if role == "admin":
            self.show_admin_controls()
        else:
            self.hide_admin_controls()

    def add_ticket(self):
        if self.current_role != "admin":
            QMessageBox.warning(self, "Ошибка", "Недостаточно прав")
            return

        try:
            # Get values from UI
            airline = self.ui.airlineInput.text()
            from_city = self.ui.fromCity.text()
            to_city = self.ui.toCity.text()
            departure_date = self.ui.departDate.date().toString("yyyy-MM-dd")
            departure_time = self.ui.departureTime.time().toString("HH:mm")
            arrival_time = self.ui.arrivalTime.time().toString("HH:mm")

            # Validate price input
            try:
                price = float(self.ui.priceInput.text())
            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Введите корректную цену")
                return

            # Validate required fields
            if not all([airline, from_city, to_city, departure_date, departure_time, arrival_time]):
                QMessageBox.warning(self, "Ошибка", "Заполните все поля")
                return

            # Add to database
            departure_datetime = f"{departure_date} {departure_time}"
            arrival_datetime = f"{departure_date} {arrival_time}"

            success = self.db.add_flight(airline, from_city, to_city,
                                       departure_datetime, arrival_datetime,
                                       price)

            if success:
                QMessageBox.information(self, "Успех", "Билет успешно добавлен")
                # Clear inputs
                self.ui.airlineInput.clear()
                self.ui.priceInput.clear()
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось добавить билет")

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def search_tickets(self):
        # Get search parameters
        from_city = self.ui.fromCity.text()
        to_city = self.ui.toCity.text()
        depart_date = self.ui.departDate.date().toString("yyyy-MM-dd")
        return_date = self.ui.returnDate.date().toString("yyyy-MM-dd")
        passengers = int(self.ui.passengersCount.currentText())

        # Debug prints
        print(f"Searching for flights:")
        print(f"From: {from_city}")
        print(f"To: {to_city}")
        print(f"Departure: {depart_date}")
        print(f"Return: {return_date}")

        # Convert class names to database format
        class_map = {
            "Эконом": "Economy",
            "Бизнес": "Business",
            "Первый": "First"
        }
        travel_class = class_map[self.ui.travelClass.currentText()]
        print(f"Class: {travel_class}")

        # Search for tickets
        outbound_tickets, return_tickets = self.db.search_tickets(
            from_city, to_city, depart_date, return_date,
            passengers, travel_class
        )

        # Debug prints
        print(f"Found outbound tickets: {len(outbound_tickets)}")
        print(f"Found return tickets: {len(return_tickets)}")

        # Show results
        self.ui.resultsLabel.show()
        self.display_tickets(outbound_tickets, return_tickets)

    def display_tickets(self, outbound_tickets, return_tickets):
        # Clear previous results
        self.ui.resultsTable.clear()
        self.ui.resultsTable.setRowCount(0)

        # Set up table headers
        headers = ["Авиакомпания", "Откуда", "Куда", "Дата вылета", "Время вылета", "Время прилета", "Цена", "Действия"]
        self.ui.resultsTable.setColumnCount(len(headers))
        self.ui.resultsTable.setHorizontalHeaderLabels(headers)
        self.ui.resultsTable.verticalHeader().setVisible(False)

        # Create a dictionary to store unique flights per airline
        unique_flights = {}

        # Process outbound tickets
        if outbound_tickets:
            for ticket in outbound_tickets:
                airline = ticket[1]
                if airline not in unique_flights:
                    unique_flights[airline] = ticket

        # Display unique flights
        if unique_flights:
            row_count = len(unique_flights)
            self.ui.resultsTable.setRowCount(row_count)

            for i, (airline, ticket) in enumerate(unique_flights.items()):
                airline_item = QTableWidgetItem(airline)
                origin = QTableWidgetItem(f"{ticket[3]} ({ticket[4]})")
                dest = QTableWidgetItem(f"{ticket[5]} ({ticket[6]})")
                date = QTableWidgetItem(ticket[7])
                dep_time = QTableWidgetItem(ticket[8])
                arr_time = QTableWidgetItem(ticket[9])
                price = QTableWidgetItem(f"{ticket[10]:.2f} ₽")

                self.ui.resultsTable.setItem(i, 0, airline_item)
                self.ui.resultsTable.setItem(i, 1, origin)
                self.ui.resultsTable.setItem(i, 2, dest)
                self.ui.resultsTable.setItem(i, 3, date)
                self.ui.resultsTable.setItem(i, 4, dep_time)
                self.ui.resultsTable.setItem(i, 5, arr_time)
                self.ui.resultsTable.setItem(i, 6, price)

                # Add Buy button with fixed size
                buy_button = QPushButton("Купить")
                buy_button.setFixedSize(100, 30)  # Устанавливаем фиксированный размер кнопки
                buy_button.clicked.connect(lambda checked, t=ticket: self.buy_ticket(t))
                buy_button.setStyleSheet("""
                    QPushButton {
                        background-color: #4e4376;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 5px;
                    }
                    QPushButton:hover {
                        background-color: #5e5386;
                    }
                """)
                
                # Создаем виджет-контейнер для центрирования кнопки
                container = QWidget()
                layout = QHBoxLayout(container)
                layout.addWidget(buy_button)
                layout.setAlignment(Qt.AlignCenter)
                layout.setContentsMargins(0, 0, 0, 0)
                self.ui.resultsTable.setCellWidget(i, 7, container)

        # Настраиваем размеры таблицы
        self.ui.resultsTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Устанавливаем фиксированную ширину для колонки с кнопками
        self.ui.resultsTable.horizontalHeader().setSectionResizeMode(7, QHeaderView.Fixed)
        self.ui.resultsTable.setColumnWidth(7, 120)
        
        # Настраиваем автоматическое изменение размера строк
        self.ui.resultsTable.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.ui.resultsTable.verticalHeader().setDefaultSectionSize(40)

        # Show table
        self.ui.resultsTable.show()

    def buy_ticket(self, ticket):
        # Проверяем, вошел ли пользователь в систему
        if not hasattr(self, 'user_id') or self.user_id is None:
            QMessageBox.warning(self, "Ошибка", "Необходимо войти в систему")
            return

        flight_data = {
            'airline': ticket[1],
            'origin': f"{ticket[3]} ({ticket[4]})",
            'destination': f"{ticket[5]} ({ticket[6]})",
            'date': ticket[7],
            'departure': ticket[8],
            'arrival': ticket[9],
            'price': ticket[10],
            'passengers': int(self.ui.passengersCount.currentText())
        }
        
        dialog = PurchaseDialog(self, flight_data)
        if dialog.exec():
            passengers_info = dialog.get_passenger_info()
            success = True
            error_message = ""
            seat_info = []
            
            # Покупаем билет для каждого пассажира
            for passenger in passengers_info:
                success, message = self.db.purchase_ticket(
                    ticket[0],
                    self.user_id,
                    passenger['name'],
                    passenger['email'],
                    passenger['phone']
                )
                if success:
                    # Добавляем информацию о месте и пассажире
                    seat_number = message.split("Номер места: ")[1]
                    seat_info.append(f"Пассажир: {passenger['name']}\nМесто: {seat_number}")
                else:
                    error_message = message
                    break
            
            if success:
                # Формируем сообщение с информацией о всех купленных билетах
                success_message = "Билеты успешно куплены!\n\n"
                success_message += "\n\n".join(seat_info)
                QMessageBox.information(self, "Успех", success_message)
                self.search_tickets()  # Обновляем список билетов
            else:
                QMessageBox.warning(self, "Ошибка", f"Ошибка при покупке билетов: {error_message}")

    def export_tickets_to_excel(self):
        if not hasattr(self, 'user_id') or self.user_id is None:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, войдите в систему")
            return
            
        tickets = self.db.get_user_tickets(self.user_id)
        if not tickets:
            QMessageBox.information(self, "Информация", "У вас нет купленных билетов")
            return
            
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Мои билеты"
        
        # Заголовки
        headers = [
            "Номер билета", "Пассажир", "Email", "Телефон", "Место",
            "Дата покупки", "Авиакомпания", "Аэропорт отправления",
            "Город отправления", "Аэропорт прибытия", "Город прибытия",
            "Дата вылета", "Дата прибытия", "Цена"
        ]
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        # Применяем заголовки и стили
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Заполняем данные
        for row, ticket in enumerate(tickets, 2):
            for col, value in enumerate(ticket, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="center")
                
        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
        # Сохраняем файл
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        filename = f"tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(desktop_path, filename)
        
        try:
            wb.save(filepath)
            QMessageBox.information(self, "Успех", f"Билеты экспортированы в файл:\n{filepath}")
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось сохранить файл:\n{str(e)}")
